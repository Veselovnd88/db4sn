import openpyxl


class Serial:
    def __init__(self, number, order, article, customer, date, comment, brand, cat_dev):
        self.number = number
        self.order = order
        self.article = article
        self.customer = customer
        self.date = date
        self.comment = comment
        self.brand = brand
        self.cat_dev = cat_dev







class Readxls:
    """ Класс берет эксель таблицу по образцу через openpyxl
    парсит ее и выдает на выходе список с серийными номерами
    """

    def __init__(self, filename):
        self.filename = filename
        self.excelfile = openpyxl.load_workbook(filename, read_only=True)
        self.sheet = self.excelfile.active
        self.snlist = []

    def givemelist(self):
        print(self.sheet.max_row)  # TODO remove
        maxcounter = 10
        if self.sheet.max_row <= maxcounter:
            for i in range(1, self.sheet.max_row + 1):
                self.snlist.append(self.sheet.cell(row=i, column=1).value)
            return self.snlist
        else:
            parts = self.sheet.max_row//maxcounter
            remainder = self.sheet.max_row%maxcounter
            print(parts)
            for i in range(0, parts):
                self.snlist = []
                for k in range(maxcounter*i+1, (i+1)*maxcounter):
                    self.snlist.append(self.sheet.cell(row=k, column=1).value)

                return self.snlist

            for k in range(parts*maxcounter,maxcounter*parts+remainder+1):
                self.snlist.append(self.sheet.cell(row=k, column=1).value)
            return self.snlist # потом будет запихивать в базу данных
        # берет из экселя номера и создает список, пока временно. В итоге сделать в sql по 100 номеров
        #


b = Readxls('test.xlsx')
mylist = b.givemelist()
print(mylist)
a = Serial('115605', '2105', '801877', 'petya', '28.11.90', 'test', 'ash','gauge')
